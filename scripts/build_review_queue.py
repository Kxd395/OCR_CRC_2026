#!/usr/bin/env python3
"""
Build Review Queue - Identifies pages needing human review
Triggers:
  - Multiple selections on same question (conflict)
  - Missing selections
  - Near-threshold detections (uncertain confidence)
  - High alignment residuals
"""
import argparse
import json
from pathlib import Path
import pandas as pd
import yaml

def load_json(p): 
    p = Path(p)
    return json.loads(p.read_text(encoding="utf-8")) if p.exists() else {}

def calculate_confidence(score, threshold, near_margin=0.03):
    """
    Calculate confidence level for a detection
    Returns: "high", "medium", "low", "very_low"
    """
    distance_from_threshold = abs(score - threshold)
    
    if distance_from_threshold > 0.10:  # >10% from threshold
        return "high"
    elif distance_from_threshold > near_margin:  # >3% from threshold
        return "medium"
    elif distance_from_threshold > 0.015:  # >1.5% from threshold
        return "low"
    else:
        return "very_low"  # Within 1.5% of threshold

def main():
    ap = argparse.ArgumentParser(description="Build human review queue for uncertain detections")
    ap.add_argument("--run-dir", help="Specific run directory (default: latest)")
    ap.add_argument("--threshold", type=float, default=None, help="Detection threshold (default: from config)")
    ap.add_argument("--near", type=float, default=0.03, help="Near-threshold margin (default: 0.03 = 3%%)")
    ap.add_argument("--residual-fail-px", type=float, default=6.0, help="Alignment residual threshold (default: 6.0px)")
    args = ap.parse_args()

    # Find run directory
    if args.run_dir:
        run = Path(args.run_dir)
    else:
        artifacts = Path("artifacts")
        runs = sorted([d for d in artifacts.glob("run_*") if d.is_dir()])
        if not runs:
            print("❌ No run directories found in artifacts/")
            return
        run = runs[-1]
        print(f"Using latest run: {run.name}")

    # Load results
    results_file = run / "step4_ocr_results" / "results.json"
    alignment_file = run / "step2_alignment_and_crop" / "alignment_results.json"
    
    if not results_file.exists():
        print(f"❌ Results not found: {results_file}")
        return
    
    ocr = load_json(results_file)
    alignment = load_json(alignment_file)

    # Get threshold
    th = args.threshold
    cfg_p = Path("configs/ocr.yaml")
    if th is None and cfg_p.exists():
        cfg = yaml.safe_load(cfg_p.read_text(encoding="utf-8"))
        th = float((cfg.get("checkbox") or {}).get("fill_threshold", 0.115))
    if th is None:
        th = 0.115

    print(f"\n{'='*70}")
    print("BUILDING REVIEW QUEUE")
    print(f"{'='*70}")
    print(f"Threshold: {th*100:.1f}%")
    print(f"Near-threshold margin: ±{args.near*100:.1f}%")
    print(f"Alignment residual threshold: {args.residual_fail_px:.1f}px")
    print(f"Total pages: {len(ocr)}")

    rows = []
    raw_rows = []
    
    for page_rec in ocr:
        page_name = page_rec.get("page")
        boxes = page_rec.get("checkboxes", [])
        
        # Store all checkbox data
        for b in boxes:
            confidence = calculate_confidence(b.get("score", 0), th, args.near)
            raw_rows.append({
                "page": page_name,
                "confidence": confidence,
                **b
            })
        
        # Analyze per question
        flags = {
            "conflict": 0,
            "missing": 0, 
            "near_threshold": 0,
            "very_low_confidence": 0
        }
        
        question_details = []
        
        # Detect questions (assumes Q1_, Q2_, etc. format)
        question_ids = set()
        for b in boxes:
            box_id = b["id"]
            # Extract question number (e.g., "Q1" from "Q1_A")
            if "_" in box_id:
                question_ids.add(box_id.split("_")[0])
        
        for qi in sorted(question_ids):
            prefix = f"{qi}_"
            q_boxes = [b for b in boxes if b["id"].startswith(prefix)]
            checked = [b for b in q_boxes if b.get("checked")]
            
            issue = None
            if len(checked) > 1:
                flags["conflict"] += 1
                issue = "CONFLICT"
                question_details.append(f"{qi}: Multiple selections ({len(checked)})")
            elif len(checked) == 0:
                flags["missing"] += 1
                issue = "MISSING"
                question_details.append(f"{qi}: No selection")
            
            # Check for near-threshold (uncertain) detections
            near_any = any(abs(b.get("score", 0) - th) <= args.near for b in q_boxes)
            if near_any:
                flags["near_threshold"] += 1
                if not issue:
                    issue = "NEAR_THRESHOLD"
                    # Find the near-threshold boxes
                    near_boxes = [b for b in q_boxes if abs(b.get("score", 0) - th) <= args.near]
                    scores_str = ", ".join([f"{b['id']}={b.get('score', 0)*100:.1f}%" for b in near_boxes])
                    question_details.append(f"{qi}: Near threshold ({scores_str})")
            
            # Check for very low confidence
            very_low = [b for b in q_boxes if calculate_confidence(b.get("score", 0), th, args.near) == "very_low"]
            if very_low:
                flags["very_low_confidence"] += len(very_low)
                if not issue:
                    issue = "LOW_CONFIDENCE"
                    scores_str = ", ".join([f"{b['id']}={b.get('score', 0)*100:.1f}%" for b in very_low])
                    question_details.append(f"{qi}: Very low confidence ({scores_str})")
        
        # Get alignment quality
        # Extract page number from various formats: page_0001.png, page_0001_aligned_cropped.png
        try:
            page_num_str = page_name.replace("page_", "").split("_")[0].replace(".png", "")
            page_num = int(page_num_str)
        except (ValueError, IndexError):
            page_num = None
        
        alignment_data = None
        residual = None
        quality = None
        
        if page_num:
            for page_result in alignment.get("pages", []):
                if page_result.get("page_number") == page_num:
                    alignment_data = page_result
                    residual = alignment_data.get("mean_error_px")
                    quality = alignment_data.get("status")
                    break
        
        # Determine if page needs review
        bad_flags = []
        if flags["conflict"] > 0:
            bad_flags.append("conflict")
        if flags["missing"] > 0:
            bad_flags.append("missing")
        if flags["near_threshold"] > 0:
            bad_flags.append("near-threshold")
        if flags["very_low_confidence"] > 0:
            bad_flags.append("low-confidence")
        if residual is not None and residual > args.residual_fail_px:
            bad_flags.append("high-residual")
        
        if bad_flags:
            priority = "HIGH" if "conflict" in bad_flags else "MEDIUM" if "near-threshold" in bad_flags or "low-confidence" in bad_flags else "LOW"
            
            rows.append({
                "priority": priority,
                "page": page_name,
                "quality": quality,
                "residual_px": residual,
                "conflicts": flags["conflict"],
                "missing": flags["missing"],
                "near_threshold": flags["near_threshold"],
                "low_confidence": flags["very_low_confidence"],
                "issues": ", ".join(bad_flags),
                "details": "; ".join(question_details),
                "recommended_action": "Manual review required"
            })

    # Create DataFrames
    dfq = pd.DataFrame(rows)
    dfraw = pd.DataFrame(raw_rows)
    
    # Add manual grading columns to review queue
    if not dfq.empty:
        # Sort by priority
        priority_order = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}
        dfq["_sort"] = dfq["priority"].map(priority_order)
        dfq = dfq.sort_values("_sort").drop("_sort", axis=1)
        
        # Add empty columns for manual grading
        dfq["reviewed"] = ""
        dfq["ocr_accurate"] = ""
        dfq["corrected_conflicts"] = ""
        dfq["reviewer_name"] = ""
        dfq["review_date"] = ""
        dfq["review_notes"] = ""

    # Output
    outdir = run / "review"
    outdir.mkdir(parents=True, exist_ok=True)
    
    dfq.to_csv(outdir / "review_queue.csv", index=False)
    
    with pd.ExcelWriter(outdir / "review_queue.xlsx", engine="openpyxl") as xl:
        if not dfq.empty:
            dfq.to_excel(xl, index=False, sheet_name="Queue")
            
            # Add formatting and data validation
            from openpyxl.worksheet.datavalidation import DataValidation
            wb = xl.book
            ws = wb["Queue"]
            
            # Make headers bold and colored
            from openpyxl.styles import Font, PatternFill
            for col in range(1, len(dfq.columns) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                # OCR columns in blue, manual grading columns in orange
                if col <= 11:  # OCR data columns
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                else:  # Manual grading columns
                    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            
            # Add data validation dropdowns for manual columns
            # "reviewed" column
            reviewed_col = chr(64 + dfq.columns.get_loc("reviewed") + 1)
            dv_reviewed = DataValidation(type="list", formula1='"YES,NO"', allow_blank=True)
            dv_reviewed.add(f'{reviewed_col}2:{reviewed_col}{len(dfq)+1}')
            ws.add_data_validation(dv_reviewed)
            
            # "ocr_accurate" column
            accurate_col = chr(64 + dfq.columns.get_loc("ocr_accurate") + 1)
            dv_accurate = DataValidation(type="list", formula1='"CORRECT,INCORRECT,PARTIAL"', allow_blank=True)
            dv_accurate.add(f'{accurate_col}2:{accurate_col}{len(dfq)+1}')
            ws.add_data_validation(dv_accurate)
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 10  # priority
            ws.column_dimensions['B'].width = 25  # page
            ws.column_dimensions['C'].width = 10  # quality
            ws.column_dimensions['D'].width = 12  # residual_px
            ws.column_dimensions['K'].width = 40  # details
            ws.column_dimensions[chr(64 + dfq.columns.get_loc("reviewed") + 1)].width = 12
            ws.column_dimensions[chr(64 + dfq.columns.get_loc("ocr_accurate") + 1)].width = 15
            ws.column_dimensions[chr(64 + dfq.columns.get_loc("corrected_conflicts") + 1)].width = 30
            ws.column_dimensions[chr(64 + dfq.columns.get_loc("reviewer_name") + 1)].width = 15
            ws.column_dimensions[chr(64 + dfq.columns.get_loc("review_date") + 1)].width = 12
            ws.column_dimensions[chr(64 + dfq.columns.get_loc("review_notes") + 1)].width = 40
        else:
            # Create empty sheet with message
            pd.DataFrame([{"message": "No pages require review - all detections are confident!"}]).to_excel(
                xl, index=False, sheet_name="Queue"
            )
        
        if not dfraw.empty:
            dfraw.to_excel(xl, index=False, sheet_name="AllCheckboxes")

    print(f"\n{'='*70}")
    print("REVIEW QUEUE SUMMARY")
    print(f"{'='*70}")
    print(f"Total pages needing review: {len(dfq)}")
    if not dfq.empty:
        print(f"\nBy priority:")
        for priority in ["HIGH", "MEDIUM", "LOW"]:
            count = len(dfq[dfq["priority"] == priority])
            if count > 0:
                print(f"  {priority}: {count} pages")
        
        print(f"\nBy issue type:")
        print(f"  Conflicts: {dfq['conflicts'].sum()}")
        print(f"  Missing: {dfq['missing'].sum()}")
        print(f"  Near threshold: {dfq['near_threshold'].sum()}")
        print(f"  Low confidence: {dfq['low_confidence'].sum()}")
    else:
        print("✅ No pages require review - all detections are confident!")
    
    print(f"\n📁 Output files:")
    print(f"  CSV: {outdir/'review_queue.csv'}")
    print(f"  Excel: {outdir/'review_queue.xlsx'}")
    print(f"{'='*70}\n")

if __name__ == "__main__":
    main()
