#!/usr/bin/env python3
"""
Export checkbox detection results to Excel for survey tallying and analysis.

This script processes all pages in a run, detects filled checkboxes using the
configured threshold, and exports the results to an Excel workbook with:
- Summary sheet with totals
- Individual page sheets
- Response distribution charts
- Raw data for further analysis

Usage:
    python scripts/export_to_excel.py <run_folder> [threshold]
    
Examples:
    python scripts/export_to_excel.py artifacts/run_20251001_185300
    python scripts/export_to_excel.py artifacts/run_20251001_185300 11.5
"""

import sys
import json
import argparse
from pathlib import Path
import cv2
import numpy as np
from datetime import datetime

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
except ImportError:
    print("ERROR: openpyxl not installed")
    print("Install with: pip install openpyxl")
    sys.exit(1)


def load_template():
    """Load the template configuration."""
    template_path = Path("templates/crc_survey_l_anchors_v1/template.json")
    with open(template_path, 'r') as f:
        return json.load(f)


def detect_checkbox_fill(roi_img):
    """
    Detect if a checkbox is filled using Otsu thresholding.
    Returns the percentage of dark pixels.
    """
    # Convert to grayscale if needed
    if len(roi_img.shape) == 3:
        gray = cv2.cvtColor(roi_img, cv2.COLOR_BGR2GRAY)
    else:
        gray = roi_img
    
    # Apply Otsu thresholding
    _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    
    # Calculate fill percentage
    total_pixels = binary.size
    dark_pixels = np.sum(binary > 0)
    fill_percent = (dark_pixels / total_pixels) * 100
    
    return fill_percent


def process_page(page_path, template, threshold):
    """
    Process a single page and detect filled checkboxes.
    Returns dict with checkbox IDs and their fill percentages.
    """
    # Load the cropped/aligned image
    img = cv2.imread(str(page_path))
    if img is None:
        return None
    
    results = {}
    
    # Process each checkbox
    for roi in template['checkbox_rois_norm']:
        checkbox_id = roi['id']
        
        # Convert normalized coordinates to pixels
        h, w = img.shape[:2]
        x = int(roi['x'] * w)
        y = int(roi['y'] * h)
        box_w = int(roi['w'] * w)
        box_h = int(roi['h'] * h)
        
        # Extract ROI
        roi_img = img[y:y+box_h, x:x+box_w]
        
        if roi_img.size == 0:
            continue
        
        # Detect fill percentage
        fill_percent = detect_checkbox_fill(roi_img)
        
        # Determine if checked
        is_checked = fill_percent >= threshold
        
        results[checkbox_id] = {
            'fill_percent': round(fill_percent, 2),
            'checked': is_checked
        }
    
    return results


def create_excel_report(run_folder, threshold, output_path=None):
    """Create comprehensive Excel report of checkbox detection results."""
    
    run_path = Path(run_folder)
    aligned_dir = run_path / "step2_alignment_and_crop" / "aligned_cropped"
    
    if not aligned_dir.exists():
        print(f"ERROR: Aligned images directory not found: {aligned_dir}")
        return None
    
    # Load template
    print("Loading template...")
    template = load_template()
    
    # Get detection settings from template
    if 'detection_settings' in template:
        template_threshold = template['detection_settings'].get('fill_threshold_percent', threshold)
        if threshold is None:
            threshold = template_threshold
    
    print(f"Using threshold: {threshold}%")
    
    # Find all aligned pages
    pages = sorted(aligned_dir.glob("page_*.png"))
    print(f"Found {len(pages)} pages to process")
    
    # Process all pages
    all_results = {}
    for page_path in pages:
        page_num = page_path.stem.split('_')[1]
        print(f"Processing page {page_num}...", end=' ')
        
        results = process_page(page_path, template, threshold)
        if results:
            all_results[page_num] = results
            checked_count = sum(1 for r in results.values() if r['checked'])
            print(f"✓ ({checked_count} checked)")
        else:
            print("✗ (failed)")
    
    if not all_results:
        print("ERROR: No results to export")
        return None
    
    # Create Excel workbook
    print("\nCreating Excel workbook...")
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create sheets
    create_summary_sheet(wb, all_results, threshold)
    create_detailed_sheet(wb, all_results, threshold)
    create_tally_sheet(wb, all_results)
    create_raw_data_sheet(wb, all_results)
    
    # Save workbook
    if output_path:
        output_file = Path(output_path)
    else:
        output_file = run_path / f"checkbox_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    print(f"\n✅ Excel report saved: {output_file}")
    
    return output_file


def create_summary_sheet(wb, all_results, threshold):
    """Create summary sheet with overview statistics."""
    ws = wb.create_sheet("Summary", 0)
    
    # Title
    ws['A1'] = "CRC Survey - Checkbox Detection Summary"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    # Metadata
    ws['A3'] = "Detection Settings"
    ws['A3'].font = Font(bold=True)
    ws['A4'] = "Threshold:"
    ws['B4'] = f"{threshold}%"
    ws['A5'] = "Method:"
    ws['B5'] = "Otsu Thresholding"
    ws['A6'] = "Total Pages:"
    ws['B6'] = len(all_results)
    
    # Calculate statistics
    total_checkboxes = 0
    total_checked = 0
    
    for page_results in all_results.values():
        total_checkboxes += len(page_results)
        total_checked += sum(1 for r in page_results.values() if r['checked'])
    
    ws['A8'] = "Overall Statistics"
    ws['A8'].font = Font(bold=True)
    ws['A9'] = "Total Checkboxes:"
    ws['B9'] = total_checkboxes
    ws['A10'] = "Total Checked:"
    ws['B10'] = total_checked
    ws['A11'] = "Total Unchecked:"
    ws['B11'] = total_checkboxes - total_checked
    ws['A12'] = "Checked Percentage:"
    ws['B12'] = f"{(total_checked/total_checkboxes*100):.1f}%"
    
    # Per-page summary
    ws['A14'] = "Per-Page Summary"
    ws['A14'].font = Font(bold=True)
    
    headers = ['Page', 'Total Boxes', 'Checked', 'Unchecked', 'Check Rate']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=15, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    row = 16
    for page_num in sorted(all_results.keys()):
        results = all_results[page_num]
        checked = sum(1 for r in results.values() if r['checked'])
        total = len(results)
        
        ws.cell(row=row, column=1, value=f"Page {page_num}")
        ws.cell(row=row, column=2, value=total)
        ws.cell(row=row, column=3, value=checked)
        ws.cell(row=row, column=4, value=total - checked)
        ws.cell(row=row, column=5, value=f"{(checked/total*100):.1f}%")
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12


def create_detailed_sheet(wb, all_results, threshold):
    """Create detailed sheet with all checkbox data."""
    ws = wb.create_sheet("Detailed Results")
    
    # Title
    ws['A1'] = "Detailed Checkbox Results by Page"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    
    # Headers
    headers = ['Page', 'Checkbox ID', 'Row', 'Column', 'Fill %', 'Status']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    row = 4
    for page_num in sorted(all_results.keys()):
        results = all_results[page_num]
        
        for checkbox_id in sorted(results.keys()):
            data = results[checkbox_id]
            
            # Parse row and column from ID (Q1_1 -> Row 1, Col 1)
            parts = checkbox_id.split('_')
            q_row = parts[0][1:]  # Remove 'Q'
            q_col = int(parts[1])  # Already 1-based
            
            ws.cell(row=row, column=1, value=f"Page {page_num}")
            ws.cell(row=row, column=2, value=checkbox_id)
            ws.cell(row=row, column=3, value=int(q_row))
            ws.cell(row=row, column=4, value=q_col)
            ws.cell(row=row, column=5, value=data['fill_percent'])
            ws.cell(row=row, column=6, value="✓ Checked" if data['checked'] else "Empty")
            
            # Color code status
            status_cell = ws.cell(row=row, column=6)
            if data['checked']:
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_cell.font = Font(color="006100", bold=True)
            
            row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12


def create_tally_sheet(wb, all_results):
    """Create tally sheet showing response distribution."""
    ws = wb.create_sheet("Response Tally")
    
    # Title
    ws['A1'] = "Response Tally - Checkbox Selections by Question"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')
    
    # Instructions
    ws['A2'] = "This sheet shows how many times each checkbox was selected across all pages"
    ws['A2'].font = Font(italic=True)
    ws.merge_cells('A2:G2')
    
    # Headers
    ws['A4'] = "Question"
    ws['B4'] = "Column 1"
    ws['C4'] = "Column 2"
    ws['D4'] = "Column 3"
    ws['E4'] = "Column 4"
    ws['F4'] = "Column 5"
    ws['G4'] = "Total Responses"
    
    for col in range(1, 8):
        cell = ws.cell(row=4, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
    
    # Tally responses
    tally = {}
    for page_results in all_results.values():
        for checkbox_id, data in page_results.items():
            if data['checked']:
                if checkbox_id not in tally:
                    tally[checkbox_id] = 0
                tally[checkbox_id] += 1
    
    # Fill in tally data
    for row_num in range(1, 6):  # Q1 through Q5
        row = 4 + row_num
        ws.cell(row=row, column=1, value=f"Q{row_num}")
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        row_total = 0
        for col_num in range(1, 6):  # Columns 1-5 (1-based)
            checkbox_id = f"Q{row_num}_{col_num}"
            count = tally.get(checkbox_id, 0)
            ws.cell(row=row, column=1+col_num, value=count)
            ws.cell(row=row, column=1+col_num).alignment = Alignment(horizontal='center')
            row_total += count
        
        ws.cell(row=row, column=7, value=row_total)
        ws.cell(row=row, column=7).font = Font(bold=True)
        ws.cell(row=row, column=7).alignment = Alignment(horizontal='center')
    
    # Column totals
    ws['A10'] = "Total"
    ws['A10'].font = Font(bold=True)
    for col in range(2, 8):
        total = sum(ws.cell(row=r, column=col).value or 0 for r in range(5, 10))
        ws.cell(row=10, column=col, value=total)
        ws.cell(row=10, column=col).font = Font(bold=True)
        ws.cell(row=10, column=col).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws.cell(row=10, column=col).alignment = Alignment(horizontal='center')
    
    # Adjust column widths
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 14


def create_raw_data_sheet(wb, all_results):
    """Create raw data sheet for pivot tables and analysis."""
    ws = wb.create_sheet("Raw Data")
    
    # Headers
    headers = ['Page', 'Checkbox_ID', 'Question', 'Column', 'Fill_Percent', 'Is_Checked']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    row = 2
    for page_num in sorted(all_results.keys()):
        results = all_results[page_num]
        
        for checkbox_id in sorted(results.keys()):
            data = results[checkbox_id]
            
            # Parse row and column (already 1-based)
            parts = checkbox_id.split('_')
            q_row = int(parts[0][1:])
            q_col = int(parts[1])
            
            ws.cell(row=row, column=1, value=int(page_num))
            ws.cell(row=row, column=2, value=checkbox_id)
            ws.cell(row=row, column=3, value=q_row)
            ws.cell(row=row, column=4, value=q_col)
            ws.cell(row=row, column=5, value=data['fill_percent'])
            ws.cell(row=row, column=6, value=1 if data['checked'] else 0)
            
            row += 1
    
    # Adjust column widths
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 14


def main():
    parser = argparse.ArgumentParser(
        description='Export checkbox detection results to Excel',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python scripts/export_to_excel.py artifacts/run_20251001_185300
  python scripts/export_to_excel.py artifacts/run_20251001_185300 11.5
  python scripts/export_to_excel.py --run-dir artifacts/run_20251001_185300 --out report.xlsx
        """
    )
    
    # Support both positional and named arguments for compatibility
    parser.add_argument('run_folder', nargs='?', help='Path to run folder (e.g., artifacts/run_20251001_185300)')
    parser.add_argument('threshold_pos', nargs='?', type=float,
                       help='Detection threshold percentage (positional)')
    parser.add_argument('--run-dir', dest='run_dir', help='Path to run folder (named argument)')
    parser.add_argument('--out', dest='output', help='Output Excel file path')
    parser.add_argument('--threshold', type=float, help='Detection threshold percentage (default: 11.5)')
    parser.add_argument('--near', type=float, default=0.03, help='± margin for near-threshold flagging (default: 0.03)')
    
    args = parser.parse_args()
    
    # Determine run folder (named argument takes precedence)
    run_folder = args.run_dir if args.run_dir else args.run_folder
    if not run_folder:
        parser.error("run_folder or --run-dir is required")
    
    # Determine threshold (named argument takes precedence)
    threshold = args.threshold if args.threshold is not None else (args.threshold_pos if args.threshold_pos is not None else 11.5)
    
    print("="*70)
    print("CRC SURVEY - EXCEL EXPORT")
    print("="*70)
    print()
    
    output_file = create_excel_report(run_folder, threshold, args.output)
    
    if output_file:
        print()
        print("="*70)
        print("✅ SUCCESS - Excel report created")
        print(f"   File: {output_file}")
        print()
        print("Sheets included:")
        print("  1. Summary       - Overview statistics")
        print("  2. Detailed      - All checkbox results")
        print("  3. Response Tally - Answer distribution")
        print("  4. Raw Data      - For pivot tables")
        print("="*70)
        
        return 0
    else:
        print("\n❌ Failed to create Excel report")
        return 1


if __name__ == '__main__':
    sys.exit(main())
